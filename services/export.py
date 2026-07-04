import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate_excel_report(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Filtrado"

    # Actualizamos los encabezados para coincidir con la nueva estructura
    headers = [
        "Fecha", 
        "Folio", 
        "Folio Pesada", 
        "Proveedor", 
        "Cantidad", 
        "Descripción", 
        "Peso Neto (kg)", 
        "Temperatura (°C)", 
        "Observación",
        "Status"
    ]
    ws.append(headers)

    # Estilos de encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        # Ajuste de ancho de columna
        ws.column_dimensions[cell.column_letter].width = 18

    # Ajustamos el ancho de la columna Descripción (ahora es la columna F)
    ws.column_dimensions['F'].width = 40 

    # Llenado de datos (Ahora acceden a las llaves del diccionario)
    for record in records:
        ws.append([
            record["Fecha"],              # Ya viene formateado como texto desde history.py
            record["Folio"],
            record["Folio_pesada"],
            record["Proveedor"],
            int(record["Cantidad"]),      # Exportado como numérico
            record["Descripcion"],
            float(record["Peso_Neto"]),   # Exportado como numérico
            float(record["Temperatura"]), # Exportado como numérico
            record["Observacion"],
            record["Status"]
        ])

    # Aplicar Autofiltro
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output